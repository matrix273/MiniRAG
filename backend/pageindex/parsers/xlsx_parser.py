"""Excel (.xlsx) 文件解析器。"""
from pathlib import Path

from openpyxl import load_workbook

from . import ParseResult


def _table_to_markdown(headers: list[str], rows: list[list[str]]) -> str:
    """将表头和数据行转为 Markdown 表格。"""
    header = "| " + " | ".join(headers) + " |"
    separator = "| " + " | ".join("---" for _ in headers) + " |"
    body_lines = ["| " + " | ".join(row) + " |" for row in rows]
    return "\n".join([header, separator] + body_lines)


def parse_xlsx(file_path: Path) -> ParseResult:
    """解析 .xlsx 文件，每个 Sheet 作为一个 section。"""
    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    content_parts: list[str] = []
    sections: list[dict] = []
    sheet_count = 0

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        # 跳过隐藏 Sheet
        if ws.sheet_state != "visible":
            continue

        all_rows: list[list] = []
        for row in ws.iter_rows(values_only=True):
            # 跳过完全为空的行
            if any(cell is not None for cell in row):
                all_rows.append(list(row))

        # 跳过空 Sheet
        if not all_rows:
            continue

        sheet_count += 1
        headers = [str(c) if c is not None else "" for c in all_rows[0]]
        data_rows = [
            [str(c) if c is not None else "" for c in row]
            for row in all_rows[1:]
        ]

        content_parts.append(f"## {ws_name}")
        content_parts.append(_table_to_markdown(headers, data_rows))

        sections.append({
            "level": 2,
            "title": ws_name,
            "text": "\n".join(str(c) for row in all_rows for c in row if c is not None),
        })

    wb.close()
    content = "\n\n".join(content_parts)
    metadata = {"sheet_count": sheet_count, "sheet_names": list(wb.sheetnames)}

    return ParseResult(content=content, metadata=metadata, sections=sections)
