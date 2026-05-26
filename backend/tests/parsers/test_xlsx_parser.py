import tempfile
from pathlib import Path
from openpyxl import Workbook

from pageindex.parsers import ParseResult
from pageindex.parsers.xlsx_parser import parse_xlsx


def _make_xlsx(sheets: dict[str, list[list]]) -> Path:
    """创建测试用 xlsx。sheets: {sheet_name: [[row1], [row2], ...]}"""
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return Path(tmp.name)


def test_parse_xlsx_single_sheet():
    path = _make_xlsx({"Data": [["Name", "Score"], ["Alice", 90], ["Bob", 85]]})
    result = parse_xlsx(path)

    assert isinstance(result, ParseResult)
    assert len(result.sections) == 1
    assert result.sections[0]["title"] == "Data"
    assert result.sections[0]["level"] == 2
    assert "## Data" in result.content
    assert "| Name |" in result.content
    assert "| Alice |" in result.content


def test_parse_xlsx_multiple_sheets():
    path = _make_xlsx({
        "Sheet1": [["A", "B"], [1, 2]],
        "Sheet2": [["X"], ["Y"]],
    })
    result = parse_xlsx(path)
    assert len(result.sections) == 2
    assert result.metadata["sheet_count"] == 2


def test_parse_xlsx_skips_empty_sheet():
    wb = Workbook()
    wb.active.title = "Empty"
    # 不添加任何数据
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    result = parse_xlsx(Path(tmp.name))
    assert len(result.sections) == 0
